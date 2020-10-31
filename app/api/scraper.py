#!/bin/usr/env python3

'''
	Scrape the pages from the following
	- Crunchbase
	- Instagram
	- Twitter
	- 

	@param: -q: query to search on angel.co

	e.g.
		python3 angel.py  -q "los angeles"
		
'''
import random
import time
import re
from datetime import datetime
import sys
import os
import selenium
import requests
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver import Firefox, Chrome, ChromeOptions, FirefoxProfile
from selenium.webdriver.firefox.options import Options as FirefoxOptions
from openpyxl import load_workbook
import pdb

from app.email import send_simple_email, send_email_with_attachment_general
from app.util import encode_csv_data
from app import create_app

logger = create_app().logger

SYSCO_USERNAME = 'cmdallm@gmail.com'
SYSCO_PASSWORD = 'UpworkTest1'
SYSCO_DATASHEET = '_Item List.xlsx'

CHENEY_USERNAME = 'cmdallm@gmail.com'
CHENEY_PASSWORD = 'cbi2644'

class Item(object):
    def __init__(self, id=None, description=None, supplier=None, order_code=None, case_size=None, case_cost=None):
        self.id = id # avoid using Python keywords where possible
        self.description = description
        self.supplier = supplier
        self.order_code = order_code
        self.case_size = case_size
        self.case_cost = case_cost

class Driver():
	basedir = os.path.abspath(os.path.dirname(__file__))
	sysco_path = f'{basedir}/data/{SYSCO_DATASHEET}'

	def __init__(self):
		self.session = requests.Session()

	def close(self):
		self.driver.close()

	def get(self, link):
		logger.info("driver: fetching link: " + link)
		status = self.get_driver().get(link)
		logger.info("driver: done fetching " + link)
		# time.sleep(20)
		return status

	def get_driver(self):
		return self.driver

	def open_chrome_browser(self):
		# Display(visible=0, size=(620, 840)).start()

		# options = ChromeOptions()
		# # options.add_argument('--user-agent=""Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/74.0.3729.157 Safari/537.36""')
		# chrome_prefs = {}
		# options.experimental_options["prefs"] = chrome_prefs
		# chrome_prefs["profile.default_content_settings"] = {"images": 2}
		# chrome_prefs["profile.managed_default_content_settings"] = {"images": 2}
		# path = f"{self.basedir}/data/chromedriver.exe"
		# chrome = Chrome(executable_path=path, options=options)
		# chrome.implicitly_wait(20)
		# self.driver = chrome

		options = FirefoxOptions()
		profile = FirefoxProfile()
		options.add_argument('--no-sandbox')
		options.add_argument('--disable-dev-shm-usage')
		options.add_argument('--headless')
		profile.set_preference("permissions.default.image", 2)
		path = f"{self.basedir}/data/geckodriver"
		firefox = Firefox(executable_path=path, options=options, firefox_profile=profile)

		self.driver = firefox

	def close_browser(self):
		logger.info('=== close browser')
		self.driver.close()

	def close_sheet(self):
		logger.info('=== close datasheet')
		self.wb.close()

	def my_send_keys(self, element, value):
		for a in value:
			time.sleep(random.randint(1,100) / 1000)
			element.send_keys(a)

	def find_sysco_item(self, item_number):
		value = ''
		el = WebDriverWait(self.driver, 10).until(EC.presence_of_element_located((By.XPATH, "//div[@class='virtualized-list']/div/div/div[contains(@class, 'price-col')]")))
		if el.text:
			value = el.text.split('\n')[0]
			logger.info(f'==== found out value {value} for {item_number}')

		return value

	def search_sysco_item(self, item_number):
		value = ''
		search_input = None
		is_done = False
		try:
			xpath = "input[data-id='myProductSearch']"
			search_input = WebDriverWait(self.driver, 30).until(EC.presence_of_element_located((By.CSS_SELECTOR, xpath)))
			self.my_send_keys(search_input, item_number)
			is_done = True
		except Exception as e:
			logger.warning(f'cannot find search element {str(e)} for {item_number}')
		try:
			time.sleep(1)
			value = self.find_sysco_item(item_number)
		except Exception as e:
			try:
				value = self.find_sysco_item(item_number)
			except Exception as e:
				logger.warning(f'cannot find price element {str(e)} for {item_number}')

		if search_input and is_done:
			search_input.clear()
		return value

	def read_sheet_as_table(self):
		self.read_datasheet()

		# header
		self.headers = []
		header_cells_generator = self.sheet.iter_rows(max_row=1)

		for cell in header_cells_generator:
			for i in range(len(cell)):
				if cell[i].value:
					self.headers.append(cell[i].value)

		# header = [cell.value for cell in self.sheet.rows[1]]
		self.items = []
		index = 0
		for row in self.sheet.rows:
			item = {
				'id': index
			}
			index += 1
			if index == 1:
				continue
			for key, cell in zip(self.headers, row):
				if cell.value:
					item[key] = cell.value

			self.items.append(item)

		return self.items, self.headers

	def read_datasheet(self):
		logger.info(' ----- Read Sheet')
		self.wb = load_workbook(filename = self.sysco_path)
		self.sheet = self.wb['Sheet1']

	def save_close_datasheet(self):
		logger.info('--- save and close sheet')
		try:
			self.wb.save(self.sysco_path)
		except Exception as e:
			logger.info(str(e))

	def update_sysco_datasheet(self):
		logger.warning(' ----- Update Sysco Sheet')
		# search order item for price
		is_done = False
		for row in self.sheet.rows:
			is_exist = False
			code_value = ''
			for cell in row:
				if  cell.column_letter == 'B' and cell.value and cell.value.strip() == 'Sysco':
					is_exist = True
				if cell.column_letter == 'C' and cell.value:
					code_value = cell.value
			if code_value:
				value = self.search_sysco_item(str(code_value))
				time.sleep(1)
				if value:
					self.sheet[f'E{cell.row}'].value = value.replace('$', '')

	def send_email(self):
		text = 'Finished the scraper. \n Please check the following url to download the excel file \n http://3.230.135.45/api/static/download'
		send_simple_email(text=text)
		# b64data = [encode_csv_data(self.sysco_path, 'Item List')]
		# html_content = '<strong>Here is the xlxs attachment after run scrapers</strong>'
		# send_email_with_attachment_general(data=b64data, html=html_content, to_email='ideveloper003@gmail.com')
		

	def scrape_sysco(self):
		self.get('https://idmsupplier.cloud.sysco.com/shop/sps/auth?PartnerId=cxs-shop-prod')
		try:
			# login
			email_field = WebDriverWait(self.driver, 10).until(EC.presence_of_element_located((By.ID, "email")))
			pwd_field = self.driver.find_element_by_id("password")
			self.my_send_keys(email_field, SYSCO_USERNAME)
			self.my_send_keys(pwd_field, SYSCO_PASSWORD)
			pwd_field.send_keys(Keys.RETURN)

			time.sleep(10)

			# go to the list page
			self.get('https://shop.sysco.com/app/lists')
			self.get('https://shop.sysco.com/app/lists')
			time.sleep(23)
		except Exception as e:
			logger.info(str(e))

	def run_scraper(self):
		self.open_chrome_browser()
	
		self.read_datasheet()

		self.run_sysco()

		self.run_cheney()

		self.close_browser()

		self.close_sheet()

		self.send_email()

	def run_sysco(self):

		self.scrape_sysco()

		self.update_sysco_datasheet()

		self.save_close_datasheet()

	def scrape_cheney(self):
		self.driver.get('https://www.cheneycentral.com/Web/#/home')
		time.sleep(10)
		try:
			# login
			login_btn = WebDriverWait(self.driver, 30).until(EC.presence_of_element_located((By.XPATH, "//a[@class='btn btn-login']")))
			login_btn.click()
			time.sleep(1)
			email_field = self.driver.find_element_by_xpath('//input[@placeholder="Email"]')
			pwd_field = self.driver.find_element_by_xpath('//input[@placeholder="Password"]')
			self.my_send_keys(email_field, CHENEY_USERNAME)
			self.my_send_keys(pwd_field, CHENEY_PASSWORD)
			pwd_field.send_keys(Keys.RETURN)

			time.sleep(10)

			# go to the list page
			self.get('https://www.cheneycentral.com/Web/#/itembook')
			self.get('https://www.cheneycentral.com/Web/#/itembook')
			time.sleep(60)
		except Exception as e:
			logger.info(str(e))

	def update_cheney_datasheet(self):
		logger.info(' ----- Update Sheet')

		# temp search
		self.search_cheney_item('234234')

		# search order item for price
		is_done = False
		for row in self.sheet.rows:
			is_exist = False
			code_value = ''
			for cell in row:
				if  cell.column_letter == 'B' and cell.value and cell.value.strip() == 'Cheney':
					is_exist = True
				if cell.column_letter == 'C' and cell.value:
					code_value = cell.value
			if code_value:
				value = self.search_cheney_item(str(code_value))
				time.sleep(1)
				if value:
					self.sheet[f'E{cell.row}'].value = value.replace('$', '')

	def find_cheney_item(self, item_number):
		value = ''
		el = WebDriverWait(self.driver, 10).until(EC.presence_of_element_located((By.XPATH, "//table[contains(@class, 'table-list')]/tbody/tr/td[6]")))
		if el.text:
			value = el.text.split('\n')[0]
			logger.info(f'==== found out value {value} for {item_number}')

		return value

	def search_cheney_item(self, item_number):
		value = ''
		search_input = None
		is_done = False
		try:
			search_input = WebDriverWait(self.driver, 30).until(EC.presence_of_element_located((By.ID, 'searchbox')))
			self.my_send_keys(search_input, item_number)
			search_input.send_keys(Keys.RETURN)
			is_done = True
		except Exception as e:
			logger.info(f'[cheney] cannot find search element {str(e)} for {item_number}')
	
		try:
			time.sleep(2)
			value = self.find_cheney_item(item_number)
		except Exception as e:
			try:
				value = self.find_cheney_item(item_number)
			except Exception as e:
				logger.info(f'[cheney] cannot find price element {str(e)} for {item_number}')

		if search_input and is_done:
			search_input.clear()
		return value


	def run_cheney(self):
		self.scrape_cheney()

		self.update_cheney_datasheet()

		self.save_close_datasheet()

if __name__ == '__main__':
	# Driver().run_sysco()
	Driver().send_email()


